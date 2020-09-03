from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import logging
import os
import datetime
import shutil
from ..models import SqliteDB
from ..common.common import varify_data_use_md5


class ExportEachObjectExcel:
    """
    export statistic data to excel
    """

    def __init__(self):
        self.book = Workbook()
        self.sqlite_db = SqliteDB()
        self.__excel_file_path_config()
        self.__each_object_sheet_init()
        self.__excel_save()

    def __excel_file_path_config(self):
        """ get excel path """
        logging.info('Start export dashborad, please waiting!!!')
        date_str = datetime.datetime.now()
        date_path = date_str.strftime("%Y-%m-%d_%H-%M-%S")
        cur_dir = os.path.dirname(os.path.abspath('__file__'))
        file_path = os.path.join(cur_dir, 'excel', date_path)

        if os.path.exists(file_path):
            shutil.rmtree(file_path)
        os.makedirs(file_path)
        self.file_path = os.path.join(
            file_path, 'oracle_objects_statistic.xls')

    def __excel_save(self):
        """ save excel """
        self.book.save(self.file_path)
        logging.info(f'bashboard excel export success, scan {self.file_path}')

    def __env_info_sheet_create(self):
        """ write env info """
        env_sheet = self.book.create_sheet('环境信息')
        self.__format_table_header(env_sheet, ['名称', '源', '目标', '校验状态'])
        env_data = self.__get_env_info_data()
        self.__write_env_info_to_excel(env_sheet, env_data)

    def __get_max_cell_width(self, cell_width_list, index, cell_value):
        """ get cell value """
        cell_width = len(str(cell_value))
        if cell_width_list[index] < cell_width:
            cell_width_list[index] = cell_width

    def __write_env_info_to_excel(self, cur_sheet, data):
        """ write env info data to excel """
        cur_row = cur_row = int(cur_sheet.max_row) + 1
        cell_width_list = []

        for row in data:
            if not cell_width_list:
                cell_width_list = [0] * len(row)

            for index, item in enumerate(row, start=1):
                self.__get_max_cell_width(cell_width_list, index-1, item)

                cur_cell = cur_sheet.cell(cur_row, index)
                cur_cell.value = item
                self.__format_normal_cell(cur_cell)
                if not row[-1]:
                    self.__format_error_cell(cur_cell)
                    self.__set_tab_color_error(cur_sheet)
            cur_row += 1

        self.__format_cell_size(cur_sheet, cell_width_list)

    def __format_cell_size(self, cur_sheet, cell_width_list):
        """ format cell size """
        alpha_list = [chr(x) for x in range(ord('A'), ord('Z') + 1)]
        for index, num in enumerate(cell_width_list, start=0):
            cur_cell_name = alpha_list[index]
            cell_width = num + 5 if num > 15 else 20
            cur_sheet.column_dimensions[cur_cell_name].width = cell_width

    def __get_env_info_data(self):
        """ get env info data """
        env_source_data = self.sqlite_db.sqlite_env_info_object_query('source')
        env_dest_data = self.sqlite_db.sqlite_env_info_object_query('dest')
        source_data = {item[0]: item for item in env_source_data}
        dest_data = {item[0]: item for item in env_dest_data}
        env_data = self.__format_env_info_data(source_data, dest_data)
        return env_data

    def __format_env_info_data(self, source, dest):
        """ format env info data """
        result = []
        for name, data in source.items():
            source_value = data[1]
            dest_value = dest[name][1]
            status = varify_data_use_md5(source_value, dest_value)
            result.append([name, source_value, dest_value, status])
        return result

    def __each_object_sheet_init(self):
        """ create excel sheet """
        self.__first_sheet_init()
        self.__env_info_sheet_create()
        self.__common_sheet_create('table')
        self.__common_sheet_create('view')
        self.__common_sheet_create('job')
        self.__common_sheet_create('synonym')
        self.__common_sheet_create('materialized_view')
        self.__common_sheet_create('trigger')
        self.__common_sheet_create('dblink')
        self.__common_sheet_create('function')
        self.__common_sheet_create('procedure')
        self.__common_sheet_create('index')
        self.__common_sheet_create('table_partition')
        self.__common_sheet_create('package')
        self.__common_sheet_create('sequence')
        self.__common_sheet_create('type')

    def __common_sheet_create(self, object_name):
        """ create each sheet data to excel """
        object_sheet = self.book.create_sheet(object_name.upper())
        self.__format_table_header(
            object_sheet, ['Schemal', '源名称', '目标名称', '校验状态'])
        object_data = self.sqlite_db.sqlite_oracle_verify_each_object_table_query(
            object_name)

        if not len(object_data):
            self.__set_tab_color_empty(object_sheet)

        self.__write_each_object_data_to_excel(object_sheet, object_data)

    def __set_tab_color_empty(self, cur_sheet):
        """ empty tab color """
        cur_sheet.sheet_properties.tabColor = 'FFFF00'

    def __set_tab_color_error(self, cur_sheet):
        """ empty tab color """
        cur_sheet.sheet_properties.tabColor = 'cc3b3b'

    def __write_each_object_data_to_excel(self, cur_sheet, data):
        """ write each object data to first sheet """
        cur_row = int(cur_sheet.max_row) + 1
        cell_width_list = []

        for row in data:
            if not cell_width_list:
                cell_width_list = [0] * len(row)

            for index, item in enumerate(row, start=1):
                self.__get_max_cell_width(cell_width_list, index-1, item)
                cur_cell = cur_sheet.cell(cur_row, index)
                cur_cell.value = item
                self.__format_normal_cell(cur_cell)
                if row[-1] == 'False':
                    self.__format_error_cell(cur_cell)
                    self.__set_tab_color_error(cur_sheet)
            cur_row += 1
        self.__format_cell_size(cur_sheet, cell_width_list)

    def __format_table_header(self, cur_sheet, header_data):
        """ format table header """
        tb_font = Font(name=u'微软雅黑', bold=True, size=11)
        for index, item in enumerate(header_data, start=1):
            cur_cell = cur_sheet.cell(1, index)
            cur_cell.value = item
            cur_cell.font = tb_font
            cur_cell.alignment = Alignment(
                horizontal='center', vertical='center')

    def __first_sheet_init(self):
        """ write first sheet """
        cur_sheet = self.book.get_sheet_by_name(self.book.sheetnames[0])
        cur_sheet.title = '概览'
        self.__format_table_header(
            cur_sheet, ['Schemal', 'Objects', '源总数', '目标总数', '校验状态'])

        objects_data = self.sqlite_db.sqlite_verify_object_statistic_query()
        self.__write_objects_statis_data_to_excel(cur_sheet, objects_data)

    def __format_normal_cell(self, cur_cell):
        """ format normal cell style """
        cell_font = Font(name=u'微软雅黑', size=11)
        cur_cell.font = cell_font
        cur_cell.alignment = Alignment(horizontal='left', vertical='center')

    def __format_error_cell(self, cur_cell):
        """ format normal cell style """
        fill = PatternFill("solid", fgColor="cc3b3b")
        cur_cell.fill = fill

    def __write_objects_statis_data_to_excel(self, cur_sheet, data):
        """ write objects statistic data to first sheet """
        cur_row = int(cur_sheet.max_row) + 1
        cell_width_list = []

        for row in data:
            if not cell_width_list:
                cell_width_list = [0] * len(row)

            for index, item in enumerate(row, start=1):
                cur_cell = cur_sheet.cell(cur_row, index)
                cur_cell.value = item
                self.__get_max_cell_width(cell_width_list, index-1, item)
                self.__format_normal_cell(cur_cell)

                if index == 2:
                    cur_cell.value = item.upper()
                if bool(row[-1]):
                    self.__format_error_cell(cur_cell)
            cur_row += 1
        self.__format_cell_size(cur_sheet, cell_width_list)
